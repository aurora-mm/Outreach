# Script for replacing the Arweave gateway in the Excel file

import os
from openpyxl import load_workbook

def replace_gateway_in_xlsx(xlsx_file, old_gateway, new_gateway):
    if not os.path.exists(xlsx_file):
        print(f"File '{xlsx_file}' does not exist.")
        return

    try:
        # Load the workbook
        workbook = load_workbook(xlsx_file)

        # Iterate over all sheets and cells
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Replace the gateway if it's in the cell value
                        cell.value = cell.value.replace(old_gateway, new_gateway)

        # Save the workbook
        workbook.save(xlsx_file)
        print(f"Updated gateway in '{xlsx_file}' to '{new_gateway}'.")

    except Exception as e:
        print(f"An error occurred: {e}")


def main():
    xlsx_file = input("Enter the path to the Excel (.xlsx) file: ")
    old_gateway = input("Enter the current gateway to replace (e.g., permagate.io): ")
    new_gateway = input("Enter the new gateway to use: ")
    replace_gateway_in_xlsx(xlsx_file, old_gateway, new_gateway)

if __name__ == "__main__":
    main()
